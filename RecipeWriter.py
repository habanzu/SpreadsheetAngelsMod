import openpyxl

get_column_letter = openpyxl.utils.cell.get_column_letter


class DataWriter(object):
    def __init__(self, wb_name):
        self.wb = openpyxl.load_workbook(wb_name)
        self.ws = self.wb.active

    def save_changes(self):
        self.wb.save('AngelsTestEdited.xlsx')


class RecipeWriter(DataWriter):
    def __init__(self, wb_name):
        super().__init__(wb_name)

    def write_recipes(self, recipes):
        """Writes recipes to a excel table. Requires a list containing recipes."""
        row = 3
        for recipe in recipes:
            resource_cells = []
            row = row + 1
            cell = self.ws.cell(row=row, column=1, value=recipe.name)
            cell._style = self.ws['A3']._style
            cell = self.ws.cell(row=row, column=2, value=1)
            cell._style = self.ws['B3']._style
            cell = self.ws.cell(row=row, column=3, value=1)
            cell._style = self.ws['C3']._style

            column = 5
            resources = list(recipe.get_resources().items())
            sorted_resources = sorted(resources, key=lambda tup: tup[1], reverse=True)
            for resource in sorted_resources:
                cell = self.ws.cell(row=row, column=column, value=resource[0].name)
                cell._style = self.ws['E3']._style
                column = column + 1
                cell = self.ws.cell(row=row, column=column, value=resource[1])
                cell._style = self.ws['F3']._style
                column = column + 1
                liter_resource_string = '={c}{r}*C{r} /B{r}'.format(c=get_column_letter(column - 1), r=row)
                cell = self.ws.cell(row=row, column=column, value=liter_resource_string)
                cell._style = self.ws['G3']._style
                column = column + 1
                reference_column = get_column_letter(column - 1)
                cell = self.ws.cell(row=row, column=column, value='=Pump / {c}{r}'.
                                    format(c=reference_column, r=row))
                cell._style = self.ws['H3']._style
                resource_cells.append(get_column_letter(column))
                column = column + 1

            string_resource_cells = (('{},'.format(row)).join(resource_cells)) + str(row)
            cell = self.ws.cell(row=row, column=4, value='=MIN({})'.format(string_resource_cells))
            cell._style = self.ws['D3']._style
